# Plot_from_Excel_auto_sets.py
# Purpose: Auto-generate N sets of XY data from Excel by blocks of rows (no manual SERIES list),
#          compute cumulative Y per set, and save all subplots into one 300 dpi PNG.

from pathlib import Path
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import itertools
import math

# === USER CONFIGURATION ===
EXCEL_PATH = "/Users/katja/PycharmProjects/Extraction_CSP/CUG25_CSP_table.xlsx"
SHEET_NAME = "Table"

# --- How sets are constructed ---
X_COL_LETTER = "K"      # Excel column letter for X (e.g., "K")
Y_COL_LETTER = "L"      # Excel column letter for Y (e.g., "L")
START_ROW = 2           # First data row of Set 1 (e.g., 2 -> K2/L2)
ROWS_PER_SET = 5        # How many rows belong to each set (your "steps" per set)
NUM_SETS = 17           # Total number of sets to include
SETS_TO_INCLUDE = []    # Can be e.g. 2, 5, 7 etc. or leave empty to include all
# If your sets are not immediately contiguous, add a gap of rows between sets:
ROWS_BETWEEN_SETS = 0   # 0 means next set starts right after previous (e.g., 2..6, 7..11, ...)

# --- Plot controls ---
X_LABEL = "Mg2+ Concentration"
Y_LABEL = "Delta_CSP (cumulative)"
SUPTITLE = "Cumulative ΔCSP for all sets"

SHOW_LINE = True
MARKER_SIZE = 30
LINE_WIDTH = 1.5
LABEL_FMT = ".4g"      # format for data labels
LABEL_OFFSET = 6       # pixels above point

# Colors: provide a list or leave [] to use Matplotlib's default cycle
COLORS = []  # e.g., ["tab:blue", "tab:orange", "tab:green", "tab:red", "tab:purple"]

# --- Output controls ---
OUTPUT_DIR = "/Users/katja/Desktop/Uno/GUC25_CSP_plots"
OUTPUT_FILE = "ScatterPlot_GUG25.png"
DPI = 300

# --- Layout controls ---
GRID_ROWS = None        # set both rows/cols to force a grid; otherwise auto
GRID_COLS = None
FIGSIZE_PER_SUBPLOT = (7, 5)  # (width, height) in inches per subplot
# === END CONFIGURATION ===


def to_numeric(seq):
    """Convert list to float, best-effort (handles strings with commas/units)."""
    out = []
    for v in seq:
        if v is None:
            out.append(math.nan)
            continue
        try:
            out.append(float(v))
        except Exception:
            s = str(v).strip().replace(",", ".")
            num = "".join(ch for ch in s if (ch.isdigit() or ch in ".-+eE"))
            try:
                out.append(float(num))
            except Exception:
                out.append(math.nan)
    return out


def read_col_range(ws, col_letter: str, start_row: int, end_row: int):
    """Read a 1-col vertical range like 'K2:K6' and return a list of values."""
    a1 = f"{col_letter}{start_row}:{col_letter}{end_row}"
    cells = ws[a1]
    return [cell.value for row in cells for cell in row]


def cumulative(series):
    """Running total of series: [y0, y0+y1, y0+y1+y2, ...] with NaN-safe behavior."""
    total = 0.0
    out = []
    for v in series:
        if pd.isna(v):
            out.append(math.nan)
        else:
            total += v
            out.append(total)
    return out


def auto_grid(n):
    """Choose a near-square rows x cols layout unless overridden."""
    if GRID_ROWS and GRID_COLS:
        return GRID_ROWS, GRID_COLS
    cols = int(math.ceil(math.sqrt(n)))
    rows = int(math.ceil(n / cols))
    return rows, cols


def main():
    excel_path = Path(EXCEL_PATH)
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    # Build the N sets programmatically (no manual SERIES list)
    series_specs = []
    step_span = ROWS_PER_SET + ROWS_BETWEEN_SETS

    # Determine which sets to include
    if SETS_TO_INCLUDE:
        chosen = [s for s in SETS_TO_INCLUDE if 1 <= s <= NUM_SETS]
    else:
        chosen = list(range(1, NUM_SETS + 1))

    for set_number in chosen:
        i = set_number - 1  # zero-based index
        start = START_ROW + i * step_span
        end = start + ROWS_PER_SET - 1
        series_specs.append({
            "name": f"Set {set_number}",
            "x_start": start,
            "x_end": end,
            "y_start": start,
            "y_end": end,
        })

    # Layout
    n = len(series_specs)
    rows, cols = auto_grid(n)
    fig_w = FIGSIZE_PER_SUBPLOT[0] * cols
    fig_h = FIGSIZE_PER_SUBPLOT[1] * rows
    fig, axes = plt.subplots(rows, cols, figsize=(fig_w, fig_h), squeeze=False)

    color_iter = itertools.cycle(COLORS if COLORS else [None])
    any_plotted = True

    for idx, spec in enumerate(series_specs):
        r = idx // cols
        c = idx % cols
        ax = axes[r][c]

        name = spec["name"]
        x_raw = read_col_range(ws, X_COL_LETTER, spec["x_start"], spec["x_end"])
        y_raw = read_col_range(ws, Y_COL_LETTER, spec["y_start"], spec["y_end"])

        x = pd.to_numeric(pd.Series(to_numeric(x_raw)), errors="coerce")
        y = pd.to_numeric(pd.Series(to_numeric(y_raw)), errors="coerce")

        # Drop pairs where either side is NaN
        mask = x.notna() & y.notna()
        x = x[mask].reset_index(drop=True)
        y = y[mask].reset_index(drop=True)

        if len(x) == 0 or len(y) == 0:
            ax.set_title(f"{name} (no numeric data)")
            ax.axis("off")
            continue

        # Make Y cumulative within this set
        y = pd.Series(cumulative(y)).astype(float)

        color = next(color_iter)
        ax.scatter(x, y, s=MARKER_SIZE, c=color if color else None)
        if SHOW_LINE:
            ax.plot(x, y, color=color if color else None, linewidth=LINE_WIDTH)

        ax.grid(True, linestyle="--", alpha=0.4)
        ax.set_xlabel(X_LABEL)
        ax.set_ylabel(Y_LABEL)
        ax.set_title(name)

        # Labels above points
        for xi, yi in zip(x, y):
            ax.annotate(f"{yi:{LABEL_FMT}}", (xi, yi),
                        textcoords="offset points", xytext=(0, LABEL_OFFSET),
                        ha="center", va="bottom")

    # Turn off any unused axes (shouldn’t happen, but safe if grid > n)
    for j in range(n, rows * cols):
        r = j // cols
        c = j % cols
        axes[r][c].axis("off")

    if SUPTITLE:
        fig.suptitle(SUPTITLE, y=0.995)

    fig.tight_layout(rect=[0, 0, 1, 0.98])

    # Save
    out_dir = Path(OUTPUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / OUTPUT_FILE
    fig.savefig(out_path, dpi=DPI, bbox_inches="tight")
    plt.close(fig)

    print(f"✅ Saved {n} subplots to: {out_path.resolve()}")


if __name__ == "__main__":
    main()