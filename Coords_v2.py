# BarColorMap_v5_cb_coords_float.py
# Adds continuous coordinate control for colorbar placement.
#
# CB_POS = (hx, vy) where each ∈ [-1, 1]
#   hx=-1 → left edge, 0→center, 1→right edge
#   vy=-1 → bottom, 0→center, 1→top
# Works with any float, e.g., (0.8, 0.8) = upper-right area.
#
# Orientation logic and all features preserved.

from pathlib import Path
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import math
import numpy as np
import re
from matplotlib.colors import (
    ListedColormap, BoundaryNorm,
    LinearSegmentedColormap, Normalize
)
from matplotlib.patches import Patch

# === USER CONFIGURATION =======================================================
EXCEL_PATH = "/Users/katja/PycharmProjects/Extraction_CSP/BEST_GUC25_Table_CSP.xlsx"
SHEET_NAME = "Table"

X_COL_LETTER = "AJ"
Y_COL_LETTER = "L"

START_ROW = 4
ROW_JUMP  = 5
NUM_SETS  = 22

SETS_TO_INCLUDE = []
INCLUDE_FIRST_N = None

X_LABEL = "1.9 mM $Mg^{2+}$ (Effective)"
Y_LABEL = "CSP (ppm)"
TITLE   = "GUC25 $Mg^{2+}$ Titrations"

EDGE_COLOR = "black"
BAR_ALPHA  = 0.95
BAR_WIDTH  = 0.85

SHOW_VALUE_LABELS = False
LABEL_FMT   = ".4g"
LABEL_OFFSET = 4
TICK_LABEL_MODE = "x_values"
SORT_BY_FIRST_NUMBER_IN_X = True
XTICK_ROTATION = 90
XTICK_HA = "center"
XTICK_PAD = 8

THRESHOLDS = [0.01, 0.02, 0.03, 0.04] # 0.03, 0.04, 0.05
COLORS     = ["deeppink", "blueviolet", "blue", "aquamarine", "yellowgreen"] # "blue""aquamarine""yellowgreen""gold"
BAR_COLORING_MODE = "continuous"
SHOW_BIN_LEGEND = True
BIN_LABELS = ["< 0.02", "0.02 – 0.04", "> 0.04"]
LEGEND_TITLE = "CSP Threshold"
SHOW_THRESHOLD_TICKS_ON_COLORBAR = True

# === COLORBAR SETTINGS ========================================================
# Continuous coordinate system: hx, vy ∈ [-1,1]
CB_POS = (0.0, 0.75)      # e.g. upper-right
CB_SIZE = (0.35, 0.05)   # width, height (fig fraction)
CB_PAD  = 0.02
CB_ORIENTATION = "auto"  # or "horizontal"/"vertical"

# Text rotation & alignment
CB_TICK_ROTATION  = 0
CB_LABEL_ROTATION = 0
CB_TICK_HA = "center"
CB_TICK_VA = "top"
CB_TICK_PAD = 6

OUTPUT_DIR  = "/Users/katja/Desktop/Uno/GUC25_CSP_plots"
OUTPUT_FILE = "CSP_GUC25_1.9mM_Mg.png"
DPI = 300
FIGSIZE = (12, 6)
# ==============================================================================


def first_int_in_string(s):
    if s is None:
        return math.inf
    m = re.search(r'(\d+)', str(s))
    return int(m.group(1)) if m else math.inf

def to_numeric(v):
    if v is None:
        return math.nan
    try:
        return float(v)
    except Exception:
        try:
            return pd.to_numeric(v, errors="coerce")
        except Exception:
            return math.nan

def read_cell(ws, col_letter, row):
    return ws[f"{col_letter}{row}"].value

def build_set_rows():
    return {i: START_ROW + (i - 1) * ROW_JUMP for i in range(1, NUM_SETS + 1)}

def choose_sets():
    if SETS_TO_INCLUDE:
        return [s for s in SETS_TO_INCLUDE if 1 <= s <= NUM_SETS]
    if INCLUDE_FIRST_N is not None:
        n = max(0, min(INCLUDE_FIRST_N, NUM_SETS))
        return list(range(1, n + 1))
    return list(range(1, NUM_SETS + 1))

def validate_thresholds(thresholds):
    if thresholds is None or len(thresholds) == 0:
        return []
    t = [float(x) for x in thresholds]
    for i in range(1, len(t)):
        if not (t[i] > t[i-1]):
            raise ValueError("THRESHOLDS must be strictly increasing.")
    return t

def make_finite_levels(y_arr, thresholds):
    finite = y_arr[np.isfinite(y_arr)]
    if finite.size == 0:
        return [0, 1]
    lo = np.nanmin(finite)
    hi = np.nanmax(finite)
    if np.isclose(lo, hi):
        lo -= 0.5; hi += 0.5
    return [lo] + thresholds + [hi]

def format_value_label(val):
    try: return format(val, LABEL_FMT)
    except Exception: return str(val)

def draw_value_labels(ax, rects, values):
    for rect, v in zip(rects, values):
        h = rect.get_height()
        ax.annotate(format_value_label(v),
                    xy=(rect.get_x() + rect.get_width()/2, h),
                    xytext=(0, LABEL_OFFSET),
                    textcoords="offset points",
                    ha="center", va="bottom", fontsize=9)

# --------------------- Colorbar coordinate placement --------------------------
def _resolve_cb_orientation(hx, vy, override):
    if override in ("horizontal", "vertical"):
        return override
    if abs(vy) > 0.1:
        return "horizontal"
    if abs(hx) > 0.1:
        return "vertical"
    return "horizontal"

def _rect_from_coords(hx, vy, size, pad):
    """hx, vy in [-1,1]; returns [x0,y0,w,h] figure fraction rect"""
    w, h = size
    # Map -1..1 to [pad, 1-pad-w] range linearly
    def interp(coord, length, size_val):
        return (coord + 1)/2 * (1 - 2*pad - size_val) + pad
    x0 = interp(hx, 1, w)
    y0 = interp(vy, 1, h)
    return [x0, y0, w, h]

def setup_colorbar_by_coords(fig, sm, thresholds, pos=(0,0), size=(0.3,0.05),
                             pad=0.02, orientation_mode="auto"):
    hx, vy = pos
    orient = _resolve_cb_orientation(hx, vy, orientation_mode)
    w, h = size
    if orient == "vertical" and w > h:
        size = (h, w)
    rect = _rect_from_coords(hx, vy, size, pad)
    cax = fig.add_axes(rect)
    cbar = fig.colorbar(sm, cax=cax, orientation=orient)

    if orient == "horizontal":
        cbar.set_label(Y_LABEL)
        cbar.ax.xaxis.label.set_rotation(CB_LABEL_ROTATION)
        cbar.ax.xaxis.set_label_position("bottom" if vy <= 0 else "top")
        if SHOW_THRESHOLD_TICKS_ON_COLORBAR and thresholds:
            cbar.set_ticks(thresholds)
        for tick in cbar.ax.get_xticklabels():
            tick.set_rotation(CB_TICK_ROTATION)
            tick.set_ha(CB_TICK_HA)
            tick.set_va(CB_TICK_VA)
        cbar.ax.tick_params(axis="x", pad=CB_TICK_PAD)
    else:
        cbar.set_label(Y_LABEL)
        cbar.ax.yaxis.label.set_rotation(90 + CB_LABEL_ROTATION)
        cbar.ax.yaxis.set_label_position("left" if hx <= 0 else "right")
        if SHOW_THRESHOLD_TICKS_ON_COLORBAR and thresholds:
            cbar.set_ticks(thresholds)
        for tick in cbar.ax.get_yticklabels():
            tick.set_rotation(0)
            tick.set_ha("left" if hx <= 0 else "right")
            tick.set_va("center")
        cbar.ax.tick_params(axis="y", pad=6)
    return cbar

def bin_indices(values, thresholds):
    idxs=[]
    for v in values:
        if not np.isfinite(v):
            idxs.append(None); continue
        i=0; placed=False
        for t in thresholds:
            if v <= t:
                idxs.append(i); placed=True; break
            i+=1
        if not placed: idxs.append(i)
    return idxs

# ------------------------------- Main ----------------------------------------
def main():
    excel_path = Path(EXCEL_PATH)
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET_NAME}' not found.")
    ws = wb[SHEET_NAME]

    set_rows = build_set_rows()
    chosen_sets = choose_sets()

    items=[]
    for s in chosen_sets:
        r=set_rows[s]
        x_raw=read_cell(ws,X_COL_LETTER,r)
        y_raw=read_cell(ws,Y_COL_LETTER,r)
        y_val=to_numeric(y_raw)
        if pd.isna(y_val): continue
        lbl=f"Set {s}" if TICK_LABEL_MODE=="set_numbers" else ("" if x_raw is None else str(x_raw))
        items.append({"set_num":s,"label":lbl,"x_raw":x_raw,"y":float(y_val),"sort_key":first_int_in_string(x_raw)})

    if not items: raise SystemExit("No valid data.")
    if TICK_LABEL_MODE=="x_values" and SORT_BY_FIRST_NUMBER_IN_X:
        items.sort(key=lambda d:(d["sort_key"],str(d["label"])))

    used_sets=[d["set_num"] for d in items]
    labels=[d["label"] for d in items]
    y_arr=np.asarray([d["y"] for d in items],dtype=float)

    t=validate_thresholds(THRESHOLDS)
    levels=make_finite_levels(y_arr,t)

    smooth_cmap=LinearSegmentedColormap.from_list("smooth_cmap",COLORS,N=256)
    smooth_norm=Normalize(vmin=np.nanmin(y_arr),vmax=np.nanmax(y_arr))
    discrete_cmap=ListedColormap(COLORS,name="disc")
    discrete_norm=BoundaryNorm(levels,ncolors=discrete_cmap.N,extend="neither")

    fig,ax=plt.subplots(figsize=FIGSIZE)
    fig.subplots_adjust(left=0.12,right=0.98,bottom=0.2,top=0.92)

    x=np.arange(len(y_arr))
    rects=ax.bar(x,y_arr,color="white",edgecolor=EDGE_COLOR,alpha=BAR_ALPHA,width=BAR_WIDTH)
    if BAR_COLORING_MODE=="continuous":
        colors=smooth_cmap(smooth_norm(y_arr))
        for r,c in zip(rects,colors): r.set_facecolor(c)
    else:
        idxs=bin_indices(y_arr,t)
        for r,i in zip(rects,idxs):
            if i is None or i<0 or i>=discrete_cmap.N: r.set_facecolor("white")
            else: r.set_facecolor(discrete_cmap(i))

    xticklabels=[f"Set {s}" for s in used_sets] if TICK_LABEL_MODE=="set_numbers" else labels
    ax.set_xticks(x,xticklabels,rotation=XTICK_ROTATION,ha=XTICK_HA)
    ax.tick_params(axis="x",pad=XTICK_PAD)
    ax.set_xlabel(X_LABEL); ax.set_ylabel(Y_LABEL); ax.set_title(TITLE)
    if SHOW_VALUE_LABELS: draw_value_labels(ax,rects,y_arr)

    sm_for_colorbar=plt.cm.ScalarMappable(cmap=smooth_cmap,norm=smooth_norm)
    sm_for_colorbar.set_array([])

    setup_colorbar_by_coords(fig,sm_for_colorbar,t,pos=CB_POS,size=CB_SIZE,pad=CB_PAD,orientation_mode=CB_ORIENTATION)

    out_dir=Path(OUTPUT_DIR); out_dir.mkdir(parents=True,exist_ok=True)
    out_path=out_dir/OUTPUT_FILE
    fig.tight_layout(); fig.savefig(out_path,dpi=DPI,bbox_inches="tight"); plt.close(fig)

    print(f"✅ Saved bar plot to: {out_path.resolve()}")
    print(f"Colorbar pos={CB_POS}, size={CB_SIZE}, pad={CB_PAD}")

if __name__=="__main__":
    main()
